"""
Little Red Coffee - Recipe Costing Data
Parsed from 'Little Red Coffee - Recipe Costing.xlsx'
Supports editing and persistence to JSON.
"""

from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Optional
import openpyxl
import json
import hashlib

EXCEL_FILE = Path(__file__).parent.parent / "inputs" / "Little Red Coffee - Recipe Costing.xlsx"
EDITS_FILE = Path(__file__).parent.parent / "data" / "ingredient_edits.json"
RECIPE_EDITS_FILE = Path(__file__).parent.parent / "data" / "recipe_edits.json"

# Sheets to skip (not recipes)
NON_RECIPE_SHEETS = {"Template", "Menu", "Alcohol", "Drink Recipes", "Costing"}


def generate_ingredient_id(name: str, category: str) -> str:
    """Generate a stable ID for an ingredient based on name and category"""
    key = f"{category}:{name}".lower()
    return hashlib.md5(key.encode()).hexdigest()[:12]


def generate_recipe_id(name: str) -> str:
    """Generate a stable ID for a recipe based on name"""
    key = f"recipe:{name}".lower()
    return hashlib.md5(key.encode()).hexdigest()[:12]


@dataclass
class Ingredient:
    """Master ingredient from the Costing sheet"""
    name: str
    category: str
    id: str = ""  # Generated from name+category
    cost: Optional[float] = None  # Total cost (e.g., $55 for 5lb bag)
    units: Optional[float] = None  # Units per package (e.g., 100 cups)
    cost_per_unit: Optional[float] = None  # Calculated cost per single unit
    unit_sale: Optional[float] = None  # Retail price per unit
    unit_profit: Optional[float] = None
    case_profit: Optional[float] = None
    supplier: Optional[str] = None
    notes: Optional[str] = None

    def __post_init__(self):
        if not self.id:
            self.id = generate_ingredient_id(self.name, self.category)


@dataclass
class RecipeIngredient:
    """An ingredient used in a recipe"""
    name: str
    quantity: Optional[float] = None
    unit_cost: Optional[float] = None
    total_cost: Optional[float] = None


@dataclass
class Recipe:
    """A recipe with costing information"""
    name: str
    id: str = ""  # Generated from name
    concept: Optional[str] = None
    submitted_by: Optional[str] = None
    portions: int = 1
    prep_time_minutes: Optional[float] = None
    proposed_sale_price: Optional[float] = None
    margin_factor: float = 1.66  # 66% margin target

    # Calculated costs
    cost_in_wages: Optional[float] = None
    cost_per_portion: Optional[float] = None
    cost_per_recipe: Optional[float] = None
    time_and_cogs: Optional[float] = None
    min_sale_price: Optional[float] = None
    profit_per_sale: Optional[float] = None

    # Ingredients
    ingredients: list[RecipeIngredient] = field(default_factory=list)

    def __post_init__(self):
        if not self.id:
            self.id = generate_recipe_id(self.name)


def parse_costing_sheet(wb: openpyxl.Workbook) -> list[Ingredient]:
    """Parse the master Costing sheet into ingredients"""
    sheet = wb["Costing"]
    ingredients = []
    current_category = "Uncategorized"

    # Skip header row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item_name = row[0]

        # Skip empty rows
        if not item_name:
            continue

        # Check if this is a category header (no cost data)
        cost = row[2]
        units = row[3]

        # Category rows have a name but no cost/units
        if cost is None and units is None:
            current_category = str(item_name).strip()
            continue

        # Parse numeric values safely
        def safe_float(val) -> Optional[float]:
            if val is None:
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        ingredient = Ingredient(
            name=str(item_name).strip(),
            category=current_category,
            cost=safe_float(row[2]),
            units=safe_float(row[3]),
            cost_per_unit=safe_float(row[4]),
            unit_sale=safe_float(row[5]),
            unit_profit=safe_float(row[6]),
            case_profit=safe_float(row[7]),
            supplier=str(row[8]).strip() if row[8] else None,
            notes=str(row[1]).strip() if row[1] else None,
        )
        ingredients.append(ingredient)

    return ingredients


def parse_recipe_sheet(sheet) -> Optional[Recipe]:
    """Parse a single recipe sheet"""
    try:
        # Read all rows at once for efficiency
        rows = list(sheet.iter_rows(min_row=1, max_row=15, values_only=True))

        if len(rows) < 3:
            return None

        def safe_float(val) -> Optional[float]:
            if val is None:
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        def safe_str(val) -> Optional[str]:
            if val is None:
                return None
            return str(val).strip()

        # Extract recipe metadata
        # Row 0: (Recipe name, [name], ..., Ingredients header...)
        # Row 1: (Concept, [concept], ..., first ingredient...)
        recipe_name = safe_str(rows[0][1]) if len(rows[0]) > 1 else sheet.title
        concept = safe_str(rows[1][1]) if len(rows) > 1 and len(rows[1]) > 1 else None
        submitted_by = safe_str(rows[2][1]) if len(rows) > 2 and len(rows[2]) > 1 else None

        # Row 3: Date, _, Number of Portions, [portions]
        portions = safe_float(rows[3][3]) if len(rows) > 3 and len(rows[3]) > 3 else 1
        portions = int(portions) if portions else 1

        # Row 4: Cuisine, _, Prep Time, [time]
        prep_time = safe_float(rows[4][3]) if len(rows) > 4 and len(rows[4]) > 3 else None

        # Row 5: Category, _, Proposed sale price, [price]
        sale_price = safe_float(rows[5][3]) if len(rows) > 5 and len(rows[5]) > 3 else None

        # Row 6: Prep area, _, Margin, [margin]
        margin = safe_float(rows[6][3]) if len(rows) > 6 and len(rows[6]) > 3 else 1.66

        # Row 7: Cost in wages, [wage_cost], Min sale price, [min_price]
        cost_wages = safe_float(rows[7][1]) if len(rows) > 7 and len(rows[7]) > 1 else None
        min_sale = safe_float(rows[7][3]) if len(rows) > 7 and len(rows[7]) > 3 else None

        # Row 8: Cost per portion, [cost]
        cost_portion = safe_float(rows[8][1]) if len(rows) > 8 and len(rows[8]) > 1 else None

        # Row 9: Cost per recipe, [cost]
        cost_recipe = safe_float(rows[9][1]) if len(rows) > 9 and len(rows[9]) > 1 else None

        # Row 10: Time and COGS, [total], Profit per sale, [profit]
        time_cogs = safe_float(rows[10][1]) if len(rows) > 10 and len(rows[10]) > 1 else None
        profit_sale = safe_float(rows[10][3]) if len(rows) > 10 and len(rows[10]) > 3 else None

        # Parse ingredients (columns F-K, starting row 2)
        # F(5): Ingredient name, G(6): Quantity, H(7): Unit Cost, K(10): Total Cost
        ingredients = []
        for row in rows[1:12]:  # Ingredients are in rows 2-12 (0-indexed: 1-11)
            if len(row) > 5 and row[5]:
                ing_name = safe_str(row[5])
                if ing_name and ing_name not in ("Ingredients", "0", "$0.00"):
                    ing = RecipeIngredient(
                        name=ing_name,
                        quantity=safe_float(row[6]) if len(row) > 6 else None,
                        unit_cost=safe_float(row[7]) if len(row) > 7 else None,
                        total_cost=safe_float(row[10]) if len(row) > 10 else None,
                    )
                    ingredients.append(ing)

        return Recipe(
            name=recipe_name or sheet.title,
            concept=concept,
            submitted_by=submitted_by,
            portions=portions,
            prep_time_minutes=prep_time,
            proposed_sale_price=sale_price,
            margin_factor=margin or 1.66,
            cost_in_wages=cost_wages,
            cost_per_portion=cost_portion,
            cost_per_recipe=cost_recipe,
            time_and_cogs=time_cogs,
            min_sale_price=min_sale,
            profit_per_sale=profit_sale,
            ingredients=ingredients,
        )
    except Exception as e:
        print(f"Error parsing recipe sheet {sheet.title}: {e}")
        return None


def load_all_data() -> tuple[list[Ingredient], list[Recipe]]:
    """Load all ingredients and recipes from the Excel file"""
    if not EXCEL_FILE.exists():
        return [], []

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Parse master ingredients
    ingredients = parse_costing_sheet(wb)

    # Parse all recipe sheets
    recipes = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in NON_RECIPE_SHEETS:
            recipe = parse_recipe_sheet(wb[sheet_name])
            if recipe:
                recipes.append(recipe)

    wb.close()
    return ingredients, recipes


# Cached data (loaded once)
_ingredients: list[Ingredient] = []
_recipes: list[Recipe] = []
_loaded = False


def get_ingredients() -> list[Ingredient]:
    """Get all master ingredients (cached)"""
    global _ingredients, _recipes, _loaded
    if not _loaded:
        _ingredients, _recipes = load_all_data()
        _apply_saved_edits()
        _apply_saved_recipe_edits()
        _loaded = True
    return _ingredients


def get_recipes() -> list[Recipe]:
    """Get all recipes (cached)"""
    global _ingredients, _recipes, _loaded
    if not _loaded:
        _ingredients, _recipes = load_all_data()
        _apply_saved_edits()
        _apply_saved_recipe_edits()
        _loaded = True
    return _recipes


def reload_data():
    """Force reload data from Excel file"""
    global _ingredients, _recipes, _loaded
    _ingredients, _recipes = load_all_data()
    _apply_saved_edits()
    _apply_saved_recipe_edits()
    _loaded = True
    return len(_ingredients), len(_recipes)


# =============================================================================
# INGREDIENT EDITING
# =============================================================================


def _load_edits() -> dict:
    """Load saved ingredient edits from JSON file"""
    if EDITS_FILE.exists():
        try:
            with open(EDITS_FILE, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def _save_edits(edits: dict):
    """Save ingredient edits to JSON file"""
    EDITS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(EDITS_FILE, "w") as f:
        json.dump(edits, f, indent=2)


def _apply_saved_edits():
    """Apply saved edits to loaded ingredients"""
    global _ingredients
    edits = _load_edits()
    if not edits:
        return

    for ing in _ingredients:
        if ing.id in edits:
            edit = edits[ing.id]
            # Apply each edited field
            for field_name, value in edit.items():
                if field_name != "id" and hasattr(ing, field_name):
                    setattr(ing, field_name, value)
            # Recalculate cost_per_unit if cost or units changed
            if ing.cost is not None and ing.units is not None and ing.units > 0:
                ing.cost_per_unit = ing.cost / ing.units
            # Recalculate unit_profit if we have the data
            if ing.unit_sale is not None and ing.cost_per_unit is not None:
                ing.unit_profit = ing.unit_sale - ing.cost_per_unit


def update_ingredient(ingredient_id: str, updates: dict) -> Optional[dict]:
    """
    Update an ingredient's fields and persist changes.
    Returns the updated ingredient dict or None if not found.
    """
    global _ingredients

    # Find the ingredient
    ing = None
    for i in _ingredients:
        if i.id == ingredient_id:
            ing = i
            break

    if not ing:
        return None

    # Load existing edits
    edits = _load_edits()
    if ingredient_id not in edits:
        edits[ingredient_id] = {}

    # Allowed fields to edit
    editable_fields = {"name", "category", "cost", "units", "unit_sale", "supplier", "notes"}

    # Apply updates
    for field_name, value in updates.items():
        if field_name in editable_fields:
            # Convert empty strings to None for numeric fields
            if field_name in {"cost", "units", "unit_sale"} and value == "":
                value = None
            elif field_name in {"cost", "units", "unit_sale"} and value is not None:
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    continue

            setattr(ing, field_name, value)
            edits[ingredient_id][field_name] = value

    # Recalculate derived values
    if ing.cost is not None and ing.units is not None and ing.units > 0:
        ing.cost_per_unit = ing.cost / ing.units
    else:
        ing.cost_per_unit = None

    if ing.unit_sale is not None and ing.cost_per_unit is not None:
        ing.unit_profit = ing.unit_sale - ing.cost_per_unit
    else:
        ing.unit_profit = None

    # Save edits
    _save_edits(edits)

    return ingredient_to_dict(ing)


def get_ingredient_by_id(ingredient_id: str) -> Optional[Ingredient]:
    """Get a single ingredient by ID"""
    for ing in get_ingredients():
        if ing.id == ingredient_id:
            return ing
    return None


def add_ingredient(data: dict) -> dict:
    """Add a new ingredient"""
    global _ingredients

    # Ensure ingredients are loaded
    get_ingredients()

    # Create new ingredient
    ing = Ingredient(
        name=data.get("name", "New Ingredient"),
        category=data.get("category", "Uncategorized"),
        cost=data.get("cost"),
        units=data.get("units"),
        unit_sale=data.get("unit_sale"),
        supplier=data.get("supplier"),
        notes=data.get("notes"),
    )

    # Calculate derived values
    if ing.cost is not None and ing.units is not None and ing.units > 0:
        ing.cost_per_unit = ing.cost / ing.units
    if ing.unit_sale is not None and ing.cost_per_unit is not None:
        ing.unit_profit = ing.unit_sale - ing.cost_per_unit

    _ingredients.append(ing)

    # Save as an edit (so it persists)
    edits = _load_edits()
    edits[ing.id] = {
        "name": ing.name,
        "category": ing.category,
        "cost": ing.cost,
        "units": ing.units,
        "unit_sale": ing.unit_sale,
        "supplier": ing.supplier,
        "notes": ing.notes,
        "_is_new": True,  # Flag to indicate this is a user-added ingredient
    }
    _save_edits(edits)

    return ingredient_to_dict(ing)


def delete_ingredient(ingredient_id: str) -> bool:
    """Delete an ingredient (mark as deleted in edits)"""
    global _ingredients

    # Find and remove from list
    for i, ing in enumerate(_ingredients):
        if ing.id == ingredient_id:
            _ingredients.pop(i)

            # Mark as deleted in edits
            edits = _load_edits()
            edits[ingredient_id] = {"_deleted": True}
            _save_edits(edits)
            return True

    return False


# =============================================================================
# RECIPE EDITING
# =============================================================================


def _load_recipe_edits() -> dict:
    """Load saved recipe edits from JSON file"""
    if RECIPE_EDITS_FILE.exists():
        try:
            with open(RECIPE_EDITS_FILE, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def _save_recipe_edits(edits: dict):
    """Save recipe edits to JSON file"""
    RECIPE_EDITS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(RECIPE_EDITS_FILE, "w") as f:
        json.dump(edits, f, indent=2)


def _apply_saved_recipe_edits():
    """Apply saved edits to loaded recipes"""
    global _recipes
    edits = _load_recipe_edits()
    if not edits:
        return

    for recipe in _recipes:
        if recipe.id in edits:
            edit = edits[recipe.id]
            # Apply each edited field
            for field_name, value in edit.items():
                if field_name not in ("id", "_deleted") and hasattr(recipe, field_name):
                    setattr(recipe, field_name, value)
            # Recalculate cost_per_portion if portions changed
            if recipe.cost_per_recipe and recipe.portions:
                recipe.cost_per_portion = recipe.cost_per_recipe / recipe.portions


def get_recipe_by_id(recipe_id: str) -> Optional[Recipe]:
    """Get a single recipe by ID"""
    for recipe in get_recipes():
        if recipe.id == recipe_id:
            return recipe
    return None


def update_recipe(recipe_id: str, updates: dict) -> Optional[dict]:
    """
    Update a recipe's fields and persist changes.
    Returns the updated recipe dict or None if not found.
    """
    global _recipes

    # Find the recipe
    recipe = None
    for r in _recipes:
        if r.id == recipe_id:
            recipe = r
            break

    if not recipe:
        return None

    # Load existing edits
    edits = _load_recipe_edits()
    if recipe_id not in edits:
        edits[recipe_id] = {}

    # Allowed fields to edit
    editable_fields = {
        "name", "concept", "portions", "prep_time_minutes",
        "proposed_sale_price", "cost_in_wages"
    }

    # Apply updates
    for field_name, value in updates.items():
        if field_name in editable_fields:
            # Convert empty strings to None for numeric fields
            numeric_fields = {"portions", "prep_time_minutes", "proposed_sale_price", "cost_in_wages"}
            if field_name in numeric_fields and value == "":
                value = None
            elif field_name in numeric_fields and value is not None:
                try:
                    if field_name == "portions":
                        value = int(value)
                    else:
                        value = float(value)
                except (ValueError, TypeError):
                    continue

            setattr(recipe, field_name, value)
            edits[recipe_id][field_name] = value

    # Recalculate derived values
    if recipe.cost_per_recipe and recipe.portions:
        recipe.cost_per_portion = recipe.cost_per_recipe / recipe.portions

    # Save edits
    _save_recipe_edits(edits)

    return recipe_to_dict(recipe)


# Helper functions for API serialization
def ingredient_to_dict(ing: Ingredient) -> dict:
    return {
        "id": ing.id,
        "name": ing.name,
        "category": ing.category,
        "cost": ing.cost,
        "units": ing.units,
        "cost_per_unit": round(ing.cost_per_unit, 4) if ing.cost_per_unit else None,
        "unit_sale": ing.unit_sale,
        "unit_profit": round(ing.unit_profit, 2) if ing.unit_profit else None,
        "case_profit": ing.case_profit,
        "supplier": ing.supplier,
        "notes": ing.notes,
    }


def recipe_to_dict(recipe: Recipe) -> dict:
    return {
        "id": recipe.id,
        "name": recipe.name,
        "concept": recipe.concept,
        "submitted_by": recipe.submitted_by,
        "portions": recipe.portions,
        "prep_time_minutes": recipe.prep_time_minutes,
        "proposed_sale_price": recipe.proposed_sale_price,
        "margin_factor": recipe.margin_factor,
        "cost_in_wages": round(recipe.cost_in_wages, 2) if recipe.cost_in_wages else None,
        "cost_per_portion": round(recipe.cost_per_portion, 2) if recipe.cost_per_portion else None,
        "cost_per_recipe": round(recipe.cost_per_recipe, 2) if recipe.cost_per_recipe else None,
        "time_and_cogs": round(recipe.time_and_cogs, 2) if recipe.time_and_cogs else None,
        "min_sale_price": round(recipe.min_sale_price, 2) if recipe.min_sale_price else None,
        "profit_per_sale": round(recipe.profit_per_sale, 2) if recipe.profit_per_sale else None,
        "ingredients": [
            {
                "name": ing.name,
                "quantity": ing.quantity,
                "unit_cost": round(ing.unit_cost, 4) if ing.unit_cost else None,
                "total_cost": round(ing.total_cost, 2) if ing.total_cost else None,
            }
            for ing in recipe.ingredients
        ],
    }


if __name__ == "__main__":
    # Test the parser
    ingredients, recipes = load_all_data()
    print(f"Loaded {len(ingredients)} ingredients and {len(recipes)} recipes")

    print("\n=== SAMPLE INGREDIENTS ===")
    for ing in ingredients[:5]:
        print(f"  {ing.name}: ${ing.cost_per_unit:.4f}/unit" if ing.cost_per_unit else f"  {ing.name}: no cost data")

    print("\n=== SAMPLE RECIPES ===")
    for recipe in recipes[:3]:
        print(f"\n  {recipe.name}")
        print(f"    Sale: ${recipe.proposed_sale_price}, Cost: ${recipe.cost_per_recipe}, Profit: ${recipe.profit_per_sale}")
        print(f"    Ingredients: {len(recipe.ingredients)}")
        for ing in recipe.ingredients[:3]:
            print(f"      - {ing.name}: qty={ing.quantity}, cost=${ing.total_cost}")
